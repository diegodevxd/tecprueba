import openpyxl
import sys

# Set stdout to use utf-8 to avoid encoding errors in terminals that don't support special chars
sys.stdout.reconfigure(encoding='utf-8')

filename = "resistencias_check.xlsx"

def check_all_images():
    print(f"Loading {filename}...")
    try:
        wb = openpyxl.load_workbook(filename)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    print(f"{'Sheet Name':<30} | {'Has Image?':<10} | {'Count':<5}")
    print("-" * 50)
    
    missing_list = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        has_image = "NO"
        
        # Check standard openpyxl images collection
        if hasattr(ws, '_images'):
            count = len(ws._images)
            if count > 0:
                has_image = "YES"
        
        # If openpyxl sees 0, checks if there are legacy drawings (rare for xlsx but possible)
        # or specific drawing relation issues.
        
        print(f"{sheet_name:<30} | {has_image:<10} | {count:<5}")
        
        if count == 0:
            missing_list.append(sheet_name)

    print("-" * 50)
    print(f"Total Sheets: {len(wb.sheetnames)}")
    print(f"Sheets without images: {len(missing_list)}")
    if missing_list:
        print("\nList of Sheets with NO extracted images:")
        for m in missing_list:
            print(f"- {m}")
            
if __name__ == "__main__":
    check_all_images()
