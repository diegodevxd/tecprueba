import openpyxl

xlsx_path = 'excel/Accesorios.xlsx'
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print('Sheets:', wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print('\n---', sheet, '---')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i>8:
            break
        print(i, row)
    # images count
    imgs = getattr(ws, '_images', [])
    print('Images in sheet:', len(imgs))
