import pandas as pd
import json
import re
import os
import openpyxl

script_file = "script.js"
output_dir = "IMAGES"
excel_file = "Capacitores Electrolíticos.xlsx"

def get_last_id():
    try:
        with open(script_file, 'r', encoding='utf-8') as f:
            content = f.read()
            ids = re.findall(r'"id":\s*(\d+)', content)
            if ids:
                return int(ids[-1])
            return 0
    except Exception as e:
        print(f"Error reading script.js: {e}")
        return 0

def sanitize_filename(name):
    # Remove invalid characters
    return re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_")

def process_sheet(ws, pid, sheet_name):
    data = {}
    
    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    
    # Extract key-value pairs
    raw_specs = {}
    for row in rows:
        if not row or len(row) < 2: continue
        k = str(row[0]).strip()
        v = str(row[1]).strip()
        if k and v and v.lower() != 'nan':
            raw_specs[k] = v

    # Construct product
    # Name: combine "Capacitor " + value if possible, or use sheet name provided it is descriptive
    # Sheet names are "Capacitor Electrolítico de 10V", etc.
    # But inside we have "Valor / especificacion": "10 V a 220 uF"
    
    val_spec = raw_specs.get("Valor / especificacion", "")
    if val_spec:
        name = f"Capacitor Electrolítico {val_spec}"
    else:
        name = sheet_name

    price = raw_specs.get("Precio por unidad", "$7 MXN")
    if "Mxn" in price:
        price = price.replace("Mxn", " MXN")
    
    # Description
    description = ""
    # Try to find description in row 13/14 logic
    desc_found = False
    for i, row in enumerate(rows):
        if row and len(row) > 0 and row[0]:
            txt = str(row[0]).strip()
            if "Descripcion" in txt:
                # Check next row column 0
                if i+1 < len(rows) and rows[i+1][0]:
                    description = str(rows[i+1][0]).strip()
                    desc_found = True
                    break
    
    if not desc_found:
        # Fallback to constructing from specs
        parts = []
        if "Valor / especificacion" in raw_specs: parts.append(raw_specs["Valor / especificacion"])
        if "Tolerancia" in raw_specs: parts.append(f"Tolerancia {raw_specs['Tolerancia']}")
        if "Potencia / voltaje" in raw_specs: parts.append(f"Voltaje {raw_specs['Potencia / voltaje']}")
        description = ", ".join(parts)

    specs = {
        "Producto": "Capacitor Electrolítico",
        "Modelo": val_spec,
        "Voltaje": raw_specs.get("Potencia / voltaje", ""),
        "Tolerancia": raw_specs.get("Tolerancia", ""),
        "Color": raw_specs.get("Color", ""),
        "SKU": f"CAP-{pid}"
    }

    # Image handling
    # Check for embedded images
    image_path = "IMAGES/capacitor_default.jpg" # Default
    if hasattr(ws, '_images') and ws._images:
        try:
            img = ws._images[0]
            ext = "png"
            if hasattr(img, 'format') and img.format:
                ext = img.format.lower()
            if ext == 'jpeg': ext = 'jpg'
            
            safe_name = sanitize_filename(name)
            filename = f"prod_{pid}_{safe_name}.{ext}"
            filepath = os.path.join(output_dir, filename)
            
            # Save image
            with open(filepath, "wb") as f:
                f.write(img._data())
            image_path = f"IMAGES/{filename}"
        except Exception as e:
            print(f"Error extracting image for {sheet_name}: {e}")

    product = {
        "id": pid,
        "category": "componentes",
        "image": image_path,
        "name": name,
        "price": price,
        "description": description.replace("\n", " "),
        "specs": specs
    }
    return product

def main():
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    print(f"Loading {excel_file}...")
    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return

    last_id = get_last_id()
    print(f"Last ID: {last_id}")
    
    new_products = []
    pid = last_id + 1
    
    for sheet_name in wb.sheetnames:
        print(f"Processing {sheet_name}...")
        ws = wb[sheet_name]
        try:
            prod = process_sheet(ws, pid, sheet_name)
            new_products.append(prod)
            pid += 1
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")

    if new_products:
        print(f"Appending {len(new_products)} products to script.js...")
        
        json_fragment = json.dumps(new_products, indent=2, ensure_ascii=False)
        # trim [ and ] from json fragment
        # But wait, json.dumps([...]) produces `[\n  {...}\n]`. 
        # I want `, \n  {...}, ...` to insert into existing array.
        
        # Strip outer brackets
        json_inner = json_fragment.strip()
        if json_inner.startswith('['): json_inner = json_inner[1:]
        if json_inner.endswith(']'): json_inner = json_inner[:-1]
        
        # Read script.js
        with open(script_file, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Find the last closing bracket of the array
        # Assuming `const products = [ ... ];`
        idx = content.rfind('];')
        if idx != -1:
            # Check if previous char is comma, if not add it
            # Actually easier to just add a comma before my fragment
            # But need to check if list is empty? "id": 1 implies not empty.
            
            insert_str = "," + json_inner + "\n"
            new_content = content[:idx] + insert_str + content[idx:]
            
            with open(script_file, 'w', encoding='utf-8') as f:
                f.write(new_content)
            print("Done!")
        else:
            print("Could not find ]; in script.js")
    else:
        print("No products created.")

if __name__ == "__main__":
    main()
