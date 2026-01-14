import pandas as pd
import json
import re
import os
import requests
import openpyxl
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

url = "https://docs.google.com/spreadsheets/d/1fdOX6rFwmrXz79wj4X-vKi9NUHaOUo49NqT5GExIdy4/export?format=xlsx"
local_filename = "resistencias_missing.xlsx"
script_file = "script.js"
output_dir = "IMAGES"

def get_last_id_and_name():
    try:
        with open(script_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # Find the last product block
            # We can use regex to find all "id": X
            ids = re.findall(r'"id":\s*(\d+)', content)
            if ids:
                last_id = int(ids[-1])
                return last_id
            return 90
    except:
        return 90

def sanitize_filename(name):
    clean = re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_").replace("Ω", "Ohm")
    return clean

def download_file():
    print(f"Downloading {url}...")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    response = requests.get(url, verify=False, headers=headers)
    response.raise_for_status()
    with open(local_filename, 'wb') as f:
        f.write(response.content)
    print("Download complete.")

def add_missing():
    try:
        download_file()
    except Exception as e:
        print(f"Download failed: {e}")
        if not os.path.exists(local_filename):
            print("No local file found inside add_missing. Exiting.")
            return

    try:
        wb = openpyxl.load_workbook(local_filename)
        sheet_names = wb.sheetnames
    except Exception as e:
        print(f"Failed to load Excel: {e}")
        return
    
    last_id = get_last_id_and_name()
    print(f"Last ID in script.js: {last_id}")
    
    next_sheet_index = last_id - 90
    
    if next_sheet_index >= len(sheet_names):
        print("No new sheets found.")
        print(f"Next index: {next_sheet_index}, Total sheets: {len(sheet_names)}")
        return

    print(f"Processing from sheet index {next_sheet_index} ({sheet_names[next_sheet_index]}) to end.")
    
    new_products = []
    pid = last_id + 1
    
    for i in range(next_sheet_index, len(sheet_names)):
        sheet_name = sheet_names[i]
        ws = wb[sheet_name]
        
        data_rows = []
        for row in ws.iter_rows(values_only=True):
            data_rows.append(row)
            
        product = {
            "id": pid,
            "category": "componentes",
            "image": "https://placehold.co/300x200?text=Resistencia", 
            "name": sheet_name.strip(),
            "price": "$1 MXN",
            "description": ""
        }
        
        if hasattr(ws, '_images') and ws._images:
            try:
                img = ws._images[0]
                ext = "png"
                if hasattr(img, 'format') and img.format:
                    ext = img.format.lower()
                if ext == 'jpeg': ext = 'jpg'
                
                safe_name = sanitize_filename(sheet_name.strip())
                filename = f"prod_{pid}_{safe_name}.{ext}"
                filepath = os.path.join(output_dir, filename)
                
                with open(filepath, "wb") as f:
                    f.write(img._data())
                
                product['image'] = f"IMAGES/{filename}"
                print(f"Saved image: {filepath}")
            except Exception as e:
                print(f"Image extraction error for {sheet_name}: {e}")
        
        for r_idx, row in enumerate(data_rows):
            if not row or len(row) < 1: continue
            if row[0] is None: continue
            
            key = str(row[0]).strip().lower()
            val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            
            if 'precio por unidad' in key:
                if val and val.lower() != 'nan':
                     product['price'] = val.replace('Mxn', ' MXN')
            
            if 'descripcion' in key:
                if val and val.lower() != 'nan':
                    product['description'] = val.replace('\n', ' ')
                elif r_idx + 1 < len(data_rows):
                    next_val = data_rows[r_idx+1][0]
                    if next_val:
                        product['description'] = str(next_val).strip().replace('\n', ' ')
        
        if not product['description']:
             specs = []
             for row in data_rows:
                if not row or len(row) < 1: continue
                if row[0] is None: continue
                k = str(row[0]).strip().lower()
                v = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                
                if 'tipo de componente' in k: specs.append(v)
                if 'valor' in k: specs.append(f"Valor: {v}")
                if 'potencia' in k: specs.append(f"Potencia: {v}")
             if specs:
                 product['description'] = ". ".join(specs).replace('\n', ' ')

        new_products.append(product)
        pid += 1

    if new_products:
        json_fragment = json.dumps(new_products, indent=2)[1:-1]
        
        with open(script_file, 'r', encoding='utf-8') as f:
            content = f.read()
            
        idx = content.rfind('];')
        if idx != -1:
            new_content = content[:idx] + "," + json_fragment + "\n" + content[idx:]
            with open(script_file, 'w', encoding='utf-8') as f:
                f.write(new_content)
            print(f"Added {len(new_products)} new products to script.js.")
        else:
            print("Could not find array end in script.js")

if __name__ == "__main__":
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    add_missing()
