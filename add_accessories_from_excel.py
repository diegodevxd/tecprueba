import json
import re
import openpyxl
import os
from PIL import Image
import io

script_path = 'script.js'
excel_path = 'excel/Accesorios.xlsx'
images_dir = 'IMAGES'

# Read script.js and extract products array
with open(script_path, 'r', encoding='utf-8') as f:
    content = f.read()

start_marker = 'const products = '
start_idx = content.find(start_marker)
if start_idx == -1:
    raise SystemExit("Couldn't find products array in script.js")
start_idx = content.find('[', start_idx)
# find matching bracket
in_string = False
escape = False
bracket_count = 0
end_idx = -1
for i in range(start_idx, len(content)):
    c = content[i]
    if not in_string and c == '"':
        in_string = True
    elif in_string:
        if escape:
            escape = False
        elif c == '\\':
            escape = True
        elif c == '"':
            in_string = False
    if not in_string:
        if c == '[':
            bracket_count += 1
        elif c == ']':
            bracket_count -= 1
            if bracket_count == 0:
                end_idx = i+1
                break
if end_idx == -1:
    raise SystemExit('Could not find end of products array')
json_text = content[start_idx:end_idx]
# clean trailing commas
json_text = re.sub(r',\s*}', '}', json_text)
json_text = re.sub(r',\s*]', ']', json_text)
products = json.loads(json_text)
max_id = max(p.get('id',0) for p in products)
next_id = max_id + 1
print(f'Loaded {len(products)} products, next id {next_id}')

# Load Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheets = [s for s in wb.sheetnames if s.lower() != 'hoja 1']
print('Accessory sheets:', sheets)

new_products = []
for sheet in sheets:
    ws = wb[sheet]
    # Build fields
    # find rows with keys in first column
    data = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row)>1 else None
        if key:
            data[key.lower()] = val
    # Compose name: prefer 'Tipo de articulo' or sheet name
    name = data.get('tipo de articulo') or sheet
    # Compose description from Marca, Modelo, Especificación, Color, Voltaje, Presentación, Artículo, Professional
    parts = []
    for k in ['marca','modelo','especificación','características','professional','article','artículo','artículo','color','voltaje','presentación','presentación','profesional','presentación ']:
        v = data.get(k)
        if v:
            parts.append(str(v))
    # fallback to some columns
    if not parts:
        # use first few entries
        for k,v in list(data.items())[:4]:
            parts.append(str(v))
    description = '. '.join(parts)
    if description and not description.endswith('.'):
        description = description + '.'
    price = '$120 MXN'
    category = 'accesorios'
    # image extraction
    img_path = None
    imgs = getattr(ws, '_images', [])
    if imgs:
        img_obj = imgs[0]
        try:
            img_bytes = img_obj._data()
            img = Image.open(io.BytesIO(img_bytes))
            # sanitize name for filename
            safe = re.sub(r'[^0-9a-zA-Z\-_\.]', '_', str(name))
            filename = f'prod_{next_id}_{safe}.png'
            os.makedirs(images_dir, exist_ok=True)
            target = os.path.join(images_dir, filename)
            img.save(target)
            img_path = f'{images_dir}/{filename}'
        except Exception as e:
            print('Image save failed for', sheet, e)
    else:
        print('No image in sheet', sheet)
    prod = {
        'id': next_id,
        'category': category,
        'image': img_path or 'IMAGES/capacitor_default.jpg',
        'name': str(name),
        'price': price,
        'description': description or ''
    }
    new_products.append(prod)
    next_id += 1

# Append new products into products list and write back to script.js
insert_point = end_idx
# Build JSON text for new products with indentation
new_json_entries = ''
for p in new_products:
    new_json_entries += json.dumps(p, ensure_ascii=False, indent=2) + ',\n'
# remove trailing comma
if new_json_entries.endswith(',\n'):
    new_json_entries = new_json_entries[:-2] + '\n'
# Insert before closing ']' of the array: if array is not empty, add a comma then new entries
# Determine if products array currently empty
if len(products) > 0:
    # place comma before new entries
    insertion = ',\n' + new_json_entries
else:
    insertion = new_json_entries
# create new content
new_content = content[:end_idx-1] + insertion + content[end_idx-1:]
# write backup
with open(script_path + '.bak', 'w', encoding='utf-8') as f:
    f.write(content)
with open(script_path, 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f'Appended {len(new_products)} products to {script_path}')
