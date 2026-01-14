import openpyxl

local_filename = "resistencias_debug.xlsx"

def inspect_sheet():
    wb = openpyxl.load_workbook(local_filename)
    sheet_name = "Resitencia 3.6Ω" # Index 11 likely
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[11]
    
    print(f"Inspecting sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    print(f"Dir ws: {dir(ws)}")
    
    if hasattr(ws, '_images'):
        print(f"ws._images: {ws._images}")
    
    if hasattr(ws, 'drawings'):
         print(f"ws.drawings: {ws.drawings}")
         
    # Check relationships
    print(f"Relations: {ws._rels}")

if __name__ == "__main__":
    inspect_sheet()
