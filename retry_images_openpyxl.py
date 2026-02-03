import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os

xlsx_path = "Capacitores Electrolíticos.xlsx"
print(f"Loading {xlsx_path} with openpyxl...")

try:
    wb = openpyxl.load_workbook(xlsx_path)
    print(f"Loaded. Sheets: {len(wb.sheetnames)}")
    
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Try finding images via standard _images
        if hasattr(ws, '_images'):
            print(f"Sheet '{sheet_name}' has {len(ws._images)} images in _images")
            count += len(ws._images)
        else:
            print(f"Sheet '{sheet_name}' has no _images attr")
            
        # Also check drawing
        if hasattr(ws, 'drawing'):
             print(f"Sheet '{sheet_name}' drawing: {ws.drawing}")

    print(f"Total images found: {count}")

except Exception as e:
    print(f"Error: {e}")
