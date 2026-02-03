import openpyxl
import os

xlsx_path = "Capacitores Electrolíticos.xlsx"
print(f"Loading {xlsx_path} with openpyxl...")

try:
    # data_only=False is default
    wb = openpyxl.load_workbook(xlsx_path)
    print(f"Loaded. Sheets: {len(wb.sheetnames)}")
    
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        img_count = 0
        
        # Method 1: _images (hidden attribute)
        if hasattr(ws, '_images'):
            img_count = len(ws._images)
        
        print(f"Sheet '{sheet_name}': {img_count} images found via _images")
        
        if img_count > 0:
            count += img_count
            # verify we can read data
            img = ws._images[0]
            print(f"  - ImageRef: {img.ref if hasattr(img, 'ref') else 'No ref'}")
            print(f"  - Format: {img.format if hasattr(img, 'format') else '?'}")

    print(f"Total images found: {count}")

except Exception as e:
    print(f"Error: {e}")
